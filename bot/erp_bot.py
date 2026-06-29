#!/usr/bin/env python3
"""
ERP DSS Telegram Bot — El Bayadh Education Directorate
Handles /status, /build, /verify, /help commands for the Academix ERP v13.4
Integrated with Hermes infrastructure via Telegram Bot API.

Environment: BOT_TOKEN from hermes-agent .env or standalone
"""

import os
import sys
import json
import asyncio
import subprocess
import logging
from datetime import datetime
from pathlib import Path

try:
    import httpx
except ImportError:
    print("httpx not installed. Run: pip install httpx")
    sys.exit(1)

# === CONFIGURATION ===
PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = PROJECT_ROOT.parent / "ERP_v13.4.xlsm"
VERIFY_RESULTS_DIR = PROJECT_ROOT / "vbe-auto" / "results"
HERMES_ENV = PROJECT_ROOT.parent / "hermes-agent" / ".env"
ALLOWED_USERS = [6562604500]  # Kamel's Telegram ID

# Read token from Hermes .env
BOT_TOKEN = None
if HERMES_ENV.exists():
    for line in HERMES_ENV.read_text().splitlines():
        if line.startswith("TELEGRAM_BOT_TOKEN="):
            BOT_TOKEN = line.split("=", 1)[1].strip()
            break

API_BASE = f"https://api.telegram.org/bot{BOT_TOKEN}"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

# === COMMAND HANDLERS ===


async def send_message(chat_id: int, text: str):
    """Send a message via Telegram Bot API."""
    async with httpx.AsyncClient(timeout=15) as client:
        try:
            resp = await client.post(
                f"{API_BASE}/sendMessage",
                json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
            )
            if resp.status_code != 200:
                logging.error(f"Telegram API error: {resp.status_code} {resp.text}")
        except Exception as e:
            logging.error(f"Failed to send message: {e}")


async def cmd_status(chat_id: int):
    """Report current ERP status — workbook, version, last verify, Ollama."""
    lines = []
    lines.append("📊 <b>ERP Académie — Status</b>")
    lines.append(f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")

    # Version
    ver = "v13.4"
    lines.append(f"📦 Version: <b>{ver}</b>")

    # Workbook
    if WORKBOOK.exists():
        size_kb = WORKBOOK.stat().st_size / 1024
        mtime = datetime.fromtimestamp(WORKBOOK.stat().st_mtime).strftime(
            "%Y-%m-%d %H:%M"
        )
        lines.append(f"📁 Workbook: <code>ERP_v13.4.xlsm</code> ({size_kb:.0f} KB, {mtime})")
    else:
        lines.append("⚠️ Workbook not found")

    # Last verify result
    results = sorted(VERIFY_RESULTS_DIR.glob("verify_results_*.json"), reverse=True)
    if results:
        latest = results[0]
        data = json.loads(latest.read_text())
        lines.append(
            f"✅ Verify: <b>{data['Passed']}/{data['Passed'] + data['Failed']}</b> PASS"
        )
        lines.append(f"   Timestamp: {data['Timestamp']}")
    else:
        lines.append("⚠️ No verify results found")

    # Ollama status
    try:
        result = subprocess.run(
            ["ollama", "ls"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        if result.returncode == 0:
            models = [l.split()[0] for l in result.stdout.splitlines()[1:] if l.strip()]
            lines.append(f"🤖 Ollama models: {', '.join(models) if models else 'none'}")
        else:
            lines.append("🤖 Ollama: not running")
    except Exception:
        lines.append("🤖 Ollama: unreachable")

    # Git info
    try:
        result = subprocess.run(
            ["git", "log", "--oneline", "-1"],
            capture_output=True,
            text=True,
            timeout=5,
            cwd=PROJECT_ROOT,
        )
        if result.returncode == 0:
            lines.append(f"🔖 Git: <code>{result.stdout.strip()}</code>")
    except Exception:
        pass

    await send_message(chat_id, "\n".join(lines))


async def cmd_build(chat_id: int):
    """Trigger a build (via build.ps1) and report result."""
    await send_message(chat_id, "🔨 <b>Build started...</b>")

    try:
        build_script = PROJECT_ROOT / "vbe-auto" / "build.ps1"
        result = subprocess.run(
            ["pwsh", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(build_script)],
            capture_output=True,
            text=True,
            timeout=600,  # 10 min timeout
            cwd=PROJECT_ROOT,
        )
        output = result.stdout[-1500:] if len(result.stdout) > 1500 else result.stdout
        if result.returncode == 0:
            await send_message(
                chat_id,
                f"✅ <b>Build successful!</b>\n<code>{output[-500:]}</code>",
            )
        else:
            await send_message(
                chat_id,
                f"❌ <b>Build failed</b> (exit code {result.returncode})\n<code>{output[-1000:]}</code>",
            )
    except subprocess.TimeoutExpired:
        await send_message(chat_id, "⏱️ Build timed out after 10 minutes")
    except Exception as e:
        await send_message(chat_id, f"❌ Build error: {e}")


async def cmd_verify(chat_id: int):
    """Run verify and report results."""
    await send_message(chat_id, "🔍 <b>Verify started...</b>")

    try:
        verify_script = PROJECT_ROOT / "vbe-auto" / "verify.ps1"
        result = subprocess.run(
            ["pwsh", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(verify_script)],
            capture_output=True,
            text=True,
            timeout=300,  # 5 min timeout
            cwd=PROJECT_ROOT,
        )
        output = result.stdout[-1500:] if len(result.stdout) > 1500 else result.stdout
        if result.returncode == 0:
            await send_message(
                chat_id,
                f"✅ <b>Verify passed!</b>\n<code>{output[-500:]}</code>",
            )
        else:
            await send_message(
                chat_id,
                f"⚠️ <b>Verify completed with issues</b>\n<code>{output[-1000:]}</code>",
            )
    except subprocess.TimeoutExpired:
        await send_message(chat_id, "⏱️ Verify timed out after 5 minutes")
    except Exception as e:
        await send_message(chat_id, f"❌ Verify error: {e}")


async def read_workbook_data():
    """Read key data from ERP workbook using openpyxl."""
    data = {}
    try:
        from openpyxl import load_workbook
        wb_path = WORKBOOK if WORKBOOK.exists() else PROJECT_ROOT / "ERP_v13.4.xlsm"
        if not wb_path.exists():
            return {"error": "Workbook not found"}

        wb = load_workbook(str(wb_path), data_only=True, keep_vba=False)

        # ── ACCUEIL KPIs ──
        acc = wb["ACCUEIL"]
        data["stock_value"] = acc.cell(12, 2).value or 0
        data["ruptures"] = acc.cell(13, 2).value or 0
        data["urgent"] = acc.cell(14, 2).value or 0
        data["alertes"] = acc.cell(15, 2).value or 0

        # ── TABLEAU DE BORD — articles with stock status ──
        tb = wb["TABLEAU DE BORD"]
        articles = []
        row = 4
        while True:
            code = tb.cell(row, 1).value
            if not code or not str(code).startswith("ART-"):
                break
            articles.append({
                "code": str(code),
                "designation": str(tb.cell(row, 2).value or ""),
                "stock": tb.cell(row, 6).value or 0,
                "rop": tb.cell(row, 10).value or 0,
                "statut": str(tb.cell(row, 11).value or ""),
                "valeur": tb.cell(row, 9).value or 0,
            })
            row += 1
        data["articles"] = articles
        data["article_count"] = len(articles)

        # ── EOQ Parameters ──
        eoq = wb["CALCULS_EOQ"]
        data["eoq_d"] = eoq.cell(2, 4).value   # D (demand)
        data["eoq_q"] = eoq.cell(2, 8).value    # Q* (EOQ)
        data["eoq_rop"] = eoq.cell(5, 4).value  # ROP
        data["eoq_ss"] = eoq.cell(5, 8).value   # Safety Stock

        wb.close()
        return data
    except Exception as e:
        return {"error": str(e)}


async def cmd_dashboard(chat_id: int):
    """Show key ERP dashboard KPIs."""
    data = read_workbook_data()
    if "error" in data:
        await send_message(chat_id, f"❌ Erreur: {data['error']}")
        return

    lines = ["📊 <b>TABLEAU DE BORD — Résumé</b>"]
    lines.append(f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")

    # KPIs
    try:
        sv = int(data.get("stock_value", 0))
        lines.append(f"💰 Valeur Stock: <b>{sv:,} DA</b>")
    except: pass
    lines.append(f"🔴 Ruptures: <b>{data.get('ruptures', 'N/A')}</b> / {data.get('article_count', 15)}")
    lines.append(f"⚠️ Urgent: <b>{data.get('urgent', 'N/A')}</b>")
    lines.append(f"✅ Alertes: <b>{data.get('alertes', 'N/A')}</b>")
    lines.append("")

    # EOQ
    lines.append("📐 <b>Paramètres EOQ (ART-001)</b>")
    lines.append(f"   Demande (D) = {data.get('eoq_d', 'N/A')}")
    lines.append(f"   EOQ (Q*)    = {data.get('eoq_q', 'N/A')}")
    lines.append(f"   ROP          = {data.get('eoq_rop', 'N/A')}")
    lines.append(f"   Stock Secu. = {data.get('eoq_ss', 'N/A')}")
    lines.append("")
    lines.append("📋 <i>Articles: détaillez avec /alerts</i>")

    await send_message(chat_id, "\n".join(lines))


async def cmd_alerts(chat_id: int):
    """Show articles that need attention (below ROP or at 0 stock)."""
    data = read_workbook_data()
    if "error" in data:
        await send_message(chat_id, f"❌ Erreur: {data['error']}")
        return

    articles = data.get("articles", [])
    lines = ["🚨 <b>Alertes Stock</b>"]
    lines.append(f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")

    ruptures = [a for a in articles if a.get("stock", 0) <= 0]
    below_rop = [a for a in articles if 0 < a.get("stock", 0) <= a.get("rop", 0)]

    if ruptures:
        lines.append(f"🔴 <b>RUPTURES ({len(ruptures)})</b>")
        for a in ruptures:
            lines.append(f"   • {a['code']} — {a['designation'][:30]} (Stock: {a['stock']})")
        lines.append("")

    if below_rop:
        lines.append(f"⚠️ <b>STOCK &lt; ROP ({len(below_rop)})</b>")
        for a in below_rop:
            lines.append(f"   • {a['code']} — {a['designation'][:30]} (Stock: {a['stock']} < ROP: {a['rop']:.0f})")
        lines.append("")

    if not ruptures and not below_rop:
        lines.append("✅ Tous les articles sont en stock suffisant.")
    else:
        ok_count = data.get("article_count", 0) - len(ruptures) - len(below_rop)
        lines.append(f"✅ {ok_count} articles en règle sur {data.get('article_count', 0)}")

    await send_message(chat_id, "\n".join(lines))


async def cmd_articles(chat_id: int):
    """List all articles with stock level and status."""
    data = read_workbook_data()
    if "error" in data:
        await send_message(chat_id, f"❌ Erreur: {data['error']}")
        return

    articles = data.get("articles", [])
    lines = ["📋 <b>Liste des Articles</b>"]
    lines.append(f"{data.get('article_count', 0)} articles — {datetime.now().strftime('%H:%M')}")
    lines.append("")

    for a in articles:
        stock = a.get("stock", 0)
        rop = a.get("rop", 0)
        status_icon = "🔴" if stock <= 0 else "⚠️" if stock <= rop else "✅"
        lines.append(f"{status_icon} <b>{a['code']}</b> — Stock: {stock} / ROP: {rop:.0f} <i>{a['statut']}</i>")

    await send_message(chat_id, "\n".join(lines))


async def cmd_help(chat_id: int):
    """Show available commands."""
    help_text = """
<b>🤖 ERP DSS Bot — Commandes</b>

/status    — État actuel du système ERP
/dashboard — KPIs et paramètres EOQ
/alerts    — Articles en alerte (rupture, stock < ROP)
/articles  — Liste complète des articles avec stock
/build     — Lancer la compilation VBA
/verify    — Exécuter la vérification (113 checks)
/help      — Cette aide

<i>Bot connecté au projet Academix v13.4
Direction de l'Éducation — El Bayadh</i>
    """.strip()
    await send_message(chat_id, help_text)


async def process_update(update: dict):
    """Process a single Telegram update."""
    message = update.get("message", {})
    chat_id = message.get("chat", {}).get("id")
    text = message.get("text", "").strip()
    user_id = message.get("from", {}).get("id")

    if not chat_id or not text:
        return

    # Authorize
    if user_id not in ALLOWED_USERS and ALLOWED_USERS:
        logging.warning(f"Unauthorized access attempt from user {user_id}")
        await send_message(chat_id, "⛔ Non autorisé. Contactez l'administrateur.")
        return

    # Route commands
    if text == "/status":
        await cmd_status(chat_id)
    elif text == "/dashboard":
        await cmd_dashboard(chat_id)
    elif text == "/alerts":
        await cmd_alerts(chat_id)
    elif text == "/articles":
        await cmd_articles(chat_id)
    elif text == "/build":
        await cmd_build(chat_id)
    elif text == "/verify":
        await cmd_verify(chat_id)
    elif text == "/help" or text.startswith("/start"):
        await cmd_help(chat_id)
    else:
        await send_message(chat_id, "Commande inconnue. Tapez /help pour la liste.")


async def poll_loop():
    """Main polling loop for Telegram updates."""
    offset = 0
    logging.info(f"ERP Bot starting (token: {BOT_TOKEN[:8]}...{BOT_TOKEN[-4:]})")
    logging.info(f"Allowed users: {ALLOWED_USERS}")

    await send_message(ALLOWED_USERS[0], "🤖 <b>ERP DSS Bot démarré</b>\nTapez /help pour les commandes.")

    async with httpx.AsyncClient(timeout=30) as client:
        while True:
            try:
                resp = await client.get(
                    f"{API_BASE}/getUpdates",
                    params={
                        "offset": offset,
                        "timeout": 30,
                        "allowed_updates": json.dumps(["message"]),
                    },
                )
                if resp.status_code != 200:
                    logging.error(f"getUpdates error: {resp.status_code}")
                    await asyncio.sleep(5)
                    continue

                data = resp.json()
                if not data.get("ok"):
                    logging.error(f"Telegram API not OK: {data}")
                    await asyncio.sleep(5)
                    continue

                for update in data.get("result", []):
                    offset = update["update_id"] + 1
                    await process_update(update)

            except httpx.TimeoutException:
                # Normal timeout when no updates — keep polling
                pass
            except Exception as e:
                logging.error(f"Poll error: {e}")
                await asyncio.sleep(5)


def main():
    if not BOT_TOKEN:
        print("ERROR: BOT_TOKEN not found. Check TELEGRAM_BOT_TOKEN in Hermes .env")
        print(f"Looked in: {HERMES_ENV}")
        sys.exit(1)

    try:
        asyncio.run(poll_loop())
    except KeyboardInterrupt:
        logging.info("Bot stopped by user")


if __name__ == "__main__":
    main()
